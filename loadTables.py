'''Subir los datos de los archivos de excel a la base de competitors'''

from pymongo import MongoClient
import openpyxl
import os
from datetime import datetime
import time

def main():
    #  abrir la base de datos
    client = MongoClient("10.192.69.41", port = 27017 )
    db = client.supplyChain
    collection = db.Competidores

    #  hacer un for loop por las carpetas
    FOLDER = "D:/myScripts/CompetitionReport/tables"
    for f in os.listdir(FOLDER):
        wb = openpyxl.load_workbook(os.path.join(FOLDER,f))
        sh = wb.active

        #leer los archivos y cargarlos a la base.
        print "reading file: " + unicode(f)
        readFile(sh, collection)


    #  agregar los nombres de competidores
    competitor_name(collection)


def readFile(sheetObject, dbcollection):
    '''toma una hoja de openpyxl y va poblando la base de datos'''

    #  data types en las columnas
    COLFECHAS = [13,14,15,16,17,18,63]
    COLNUMEROS = [30,31,32,33,34,35,36,37,40,41,48,49,50,51,52,53,54,55]
    COLPARTIDA = [6]
    COLREFRENDO = 10

    #  leer la fila uno y obtener un diccionario de headers
    headers = {}
    for n in range(1, sheetObject.max_column+1):
        headers[n] = sheetObject.cell(row=1, column=n).value

    #por cada fila desde la 2
    for nrow in range (2, sheetObject.max_row):

        #  primero busca el refrendo, y si ya esta ingresado, pasa a la siguiente fila
        #refrendo = unicode(sheetObject.cell(row=nrow, column=COLREFRENDO).value).strip()
        #if dbcollection.find({"REFRENDO": refrendo}).count() > 0:
        #    continue


        #  contiene todos los valores de una fila
        data = {}

        #por cada columna
        for ncol in range(1, sheetObject.max_column):


            cellvalue = unicode(sheetObject.cell(row=nrow, column=ncol).value).strip()

            #si la columna es de fecha
            if ncol in COLFECHAS:
                #cellvalue lo convierto a date
                print cellvalue
                try:
                    cellvalue = datetime.strptime(cellvalue,"%Y-%m-%d")
                except ValueError:
                    cellvalue = None

            #elif si es numeros lo conviero a num
            elif ncol in COLNUMEROS:
                try:
                    cellvalue = float(cellvalue)
                except ValueError:
                    cellvalue = None

            #elif es partida
            elif ncol in COLPARTIDA:
                #al cell value quitarle los puntos y el strip
                cellvalue = cellvalue.replace(".", "").strip()

            #  inserta los valores en el diccionario de los datos de fila
            data.update({headers[ncol]:cellvalue})

        #  agregar la categoria del producto
        data.update({"PRODUCT_CATEGORY": None})
        data.update({"COMPETITOR_NAME": None})

        #insertar el diccionario en la base de datos.
        dbcollection.insert_one(data)


def competitor_name(collection):
    names = {
        "1790022404001": "ALCATEL",
        "1792040825001": "CISCO",
        "1768152560001": "CNT",
        "1791251237001": "CLARO",
        "1790224716001": "DIGITEC",
        "1791256115001": "MOVISTAR",
        "1791862295001": "NOKIA",
        "1791845722001": "HUAWEI",
        "1791846842001": "ZTE"
    }

    CUSTOMERS = ["CLARO", "MOVISTAR", "CNT"]

    null_competitors = collection.find({"COMPETITOR_NAME": None})

    #si no hay nulls, entonces termina.
    if null_competitors.count() < 1: return


    for document in null_competitors:
        if not document["RUC"] in names.keys():
            continue

        #  si es el ruc de un cliente
        if names[document["RUC"]] in CUSTOMERS:

            if "CISCO" in document["EMBARCADOR"]:
                collection.update({"_id":document["_id"]}, {"$set":{"COMPETITOR_NAME":"CISCO"}})
                #document["COMPETITOR_NAME"] = "CISCO"
            elif "NOKIA" in document["EMBARCADOR"]:
                collection.update({"_id": document["_id"]}, {"$set":{"COMPETITOR_NAME": "NOKIA"}})
                #document["COMPETITOR_NAME"] = "NOKIA"
            elif "ZTE" in document["EMBARCADOR"]:
                collection.update({"_id": document["_id"]}, {"$set":{"COMPETITOR_NAME": "ZTE"}})
                #document["COMPETITOR_NAME"] = "ZTE"
            elif "DIGITEC" in document["EMBARCADOR"]:
                collection.update({"_id": document["_id"]}, {"$set":{"COMPETITOR_NAME": "DIGITEC"}})
                #document["COMPETITOR_NAME"] = "DIGITEC"
            elif "ALCATEL" in document["EMBARCADOR"]:
                collection.update({"_id": document["_id"]}, {"$set":{"COMPETITOR_NAME": "ALCATEL"}})
                #document["COMPETITOR_NAME"] = "ALCATEL"
            elif "HUAWEI" in document["EMBARCADOR"]:
                collection.update({"_id": document["_id"]}, {"$set":{"COMPETITOR_NAME": "HUAWEI"}})
                #document["COMPETITOR_NAME"] = "HUAWEI"

        else:
            collection.update({"_id": document["_id"]}, {"$set":{"COMPETITOR_NAME": names[document["RUC"]]}})
            #document["COMPETITOR_NAME"] = names[document["RUC"]]



if __name__ == "__main__":
    main()
    # client = MongoClient("10.192.69.41", port=27017)
    # db = client.supplyChain
    # collection = db.Competidores
    # competitor_name(collection)
